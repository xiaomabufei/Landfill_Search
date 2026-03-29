"""读取原始数据文件，支持多种格式，提取填埋场基础信息。

支持格式: .xlsx, .xls, .csv, .tsv, .json
"""

from typing import List, Dict, Optional
import json
import csv
from pathlib import Path

# 输入文件列名 → 内部字段名的映射（兼容不同格式的列名）
COLUMN_MAPPING = {
    # 基础信息
    "Code1": "code", "code": "code", "id": "code", "ID": "code",
    "name": "name", "Name": "name", "landfill_name": "name",
    "lat": "lat", "latitude": "lat", "Latitude": "lat",
    "lng": "lng", "lon": "lng", "longitude": "lng", "Longitude": "lng",
    "GID_0": "country_code", "country_code": "country_code",
    "Country": "country", "country": "country",

    # 原有 7 项指标
    "landfill type": "landfill_type",
    "landfill_type": "landfill_type",
    "ref 2": "landfill_type_ref",
    "whether having landfill gas collection technology": "has_gas_collection",
    "has_gas_collection": "has_gas_collection",
    "ref 3": "has_gas_collection_ref",
    "landfill gas collection technology type": "gas_collection_technology",
    "gas_collection_technology": "gas_collection_technology",
    "ref 4": "gas_collection_technology_ref",
    "landfill gas collection rate": "gas_collection_rate",
    "gas_collection_rate": "gas_collection_rate",
    "ref 5": "gas_collection_rate_ref",
    "start year": "start_year",
    "start_year": "start_year",
    "ref 6": "start_year_ref",
    "final year": "final_year",
    "final_year": "final_year",
    "ref 7": "final_year_ref",
    "start year of landfill gas collection systems": "gas_collection_start_year",
    "gas_collection_start_year": "gas_collection_start_year",
    "ref 9": "gas_collection_start_year_ref",

    # 扩展指标 — 填埋场常用信息
    "waste_capacity_m3": "waste_capacity_m3",
    "waste capacity": "waste_capacity_m3",
    "capacity": "waste_capacity_m3",
    "annual_waste_intake_tons": "annual_waste_intake_tons",
    "waste_types_accepted": "waste_types_accepted",
    "waste types": "waste_types_accepted",
    "area_hectares": "area_hectares",
    "area": "area_hectares",
    "operator": "operator",
    "depth_meters": "depth_meters",
    "liner_type": "liner_type",
    "leachate_treatment": "leachate_treatment",
    "status": "status",
    "environmental_issues": "environmental_issues",
    "methane_emission_tons_year": "methane_emission_tons_year",
}

# 所有指标字段（用于检查完整性）
CORE_INDICATORS = [
    "landfill_type", "has_gas_collection", "gas_collection_technology",
    "gas_collection_rate", "start_year", "final_year", "gas_collection_start_year",
]

EXTENDED_INDICATORS = [
    "waste_capacity_m3", "annual_waste_intake_tons", "waste_types_accepted",
    "area_hectares", "operator", "depth_meters", "liner_type",
    "leachate_treatment", "status", "environmental_issues", "methane_emission_tons_year",
]

ALL_INDICATORS = CORE_INDICATORS + EXTENDED_INDICATORS

# 基础字段
BASE_FIELDS = ["code", "name", "lat", "lng", "country_code", "country"]


def _normalize_row(raw_data: Dict, column_map: Dict) -> Dict:
    """将原始数据行通过列名映射转换为标准格式。"""
    landfill = {f: None for f in BASE_FIELDS + ALL_INDICATORS}
    # 加上 ref 字段
    for indicator in CORE_INDICATORS:
        landfill[indicator + "_ref"] = None

    for raw_key, value in raw_data.items():
        if raw_key is None:
            continue
        mapped = column_map.get(raw_key.strip() if isinstance(raw_key, str) else raw_key)
        if mapped:
            landfill[mapped] = value

    return landfill


def _read_xlsx(file_path: str) -> List[Dict]:
    """读取 .xlsx / .xls 文件。"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(_normalize_row(dict(zip(headers, row)), COLUMN_MAPPING))
    wb.close()
    return rows


def _read_csv(file_path: str, delimiter: str = ",") -> List[Dict]:
    """读取 .csv / .tsv 文件。"""
    rows = []
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            rows.append(_normalize_row(row, COLUMN_MAPPING))
    return rows


def _read_json(file_path: str) -> List[Dict]:
    """读取 .json 文件。支持数组格式或 {landfills: [...]} 格式。"""
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        items = data
    elif isinstance(data, dict) and "landfills" in data:
        items = data["landfills"]
    else:
        raise ValueError("JSON 格式不支持：需要数组或包含 'landfills' 字段的对象")

    rows = []
    for item in items:
        rows.append(_normalize_row(item, COLUMN_MAPPING))
    return rows


def read_landfills(file_path: str) -> List[Dict]:
    """读取填埋场数据文件，自动识别格式。

    支持: .xlsx, .xls, .csv, .tsv, .json
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        return _read_xlsx(file_path)
    elif suffix == ".csv":
        return _read_csv(file_path, delimiter=",")
    elif suffix == ".tsv":
        return _read_csv(file_path, delimiter="\t")
    elif suffix == ".json":
        return _read_json(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {suffix}（支持 .xlsx, .xls, .csv, .tsv, .json）")


def get_unfilled_landfills(landfills: List[Dict]) -> List[Dict]:
    """筛选出核心指标全部为空的填埋场。"""
    unfilled = []
    for lf in landfills:
        if all(lf.get(field) is None for field in CORE_INDICATORS):
            unfilled.append(lf)
    return unfilled


def get_batches(landfills: List[Dict], batch_size: int = 10) -> List[List[Dict]]:
    """将填埋场列表按批次分组。"""
    return [landfills[i:i + batch_size] for i in range(0, len(landfills), batch_size)]


def get_data_summary(landfills: List[Dict]) -> Dict:
    """生成数据概要统计。"""
    total = len(landfills)
    if total == 0:
        return {"total": 0}

    filled = {}
    for field in ALL_INDICATORS:
        count = sum(1 for lf in landfills if lf.get(field) is not None)
        filled[field] = {"count": count, "rate": round(count / total * 100, 1)}

    return {
        "total": total,
        "unfilled": len(get_unfilled_landfills(landfills)),
        "fields": filled,
    }


if __name__ == "__main__":
    import sys
    data_path = sys.argv[1] if len(sys.argv) > 1 else str(Path(__file__).parent.parent.parent / "data" / "raw" / "ITA.xlsx")

    landfills = read_landfills(data_path)
    summary = get_data_summary(landfills)

    print(f"文件: {data_path}")
    print(f"格式: {Path(data_path).suffix}")
    print(f"总填埋场数: {summary['total']}")
    print(f"待填写数: {summary['unfilled']}")
    print(f"\n指标完成率:")
    for field, info in summary["fields"].items():
        bar = "█" * (info["count"] * 20 // summary["total"]) if summary["total"] > 0 else ""
        print(f"  {field:35s} {info['count']:>4d}/{summary['total']}  {info['rate']:>5.1f}%  {bar}")
